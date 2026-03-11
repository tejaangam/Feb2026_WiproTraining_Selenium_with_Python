from openpyxl import load_workbook

def get_login_data(path):
    workbook = load_workbook(path)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        username = row[0]
        password = row[1]
        data.append((username, password))

    return data